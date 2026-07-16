#!/usr/bin/env python3
"""Собрать HTML из programs.json.

Источник данных — только programs.json (смотри блок _как_править внутри файла).
После правок: python3 build.py
"""

from __future__ import annotations

import base64
import json
import html as H
import urllib.request
import zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent
PROGRAMS_FILE = ROOT / "programs.json"
OUT_HTML = ROOT / "Гиря.html"
OUT_ZIP = ROOT / "Гиря.zip"
OUT_XLSX_FEMALE = ROOT / "Две_тренировки_для_девушки.xlsx"
OUT_XLSX_MALE = ROOT / "Две_тренировки_для_парня.xlsx"
GIF_CACHE = ROOT / ".gif-cache"


def esc(s: object) -> str:
    return H.escape(str(s))


def load_gif_bytes(gif_value: str) -> bytes | None:
    if not gif_value:
        return None

    if gif_value.startswith("data:image"):
        header, b64 = gif_value.split(",", 1)
        return base64.b64decode(b64)

    path = Path(gif_value)
    if not path.is_absolute():
        path = ROOT / path

    if path.exists():
        return path.read_bytes()

    if gif_value.startswith(("http://", "https://")):
        cache_name = gif_value.rsplit("/", 1)[-1]
        GIF_CACHE.mkdir(exist_ok=True)
        cache_path = GIF_CACHE / cache_name
        if not cache_path.exists():
            print(f"  скачиваю {gif_value}")
            req = urllib.request.Request(gif_value, headers={"User-Agent": "Mozilla/5.0"})
            try:
                with urllib.request.urlopen(req, timeout=60) as resp:
                    cache_path.write_bytes(resp.read())
            except Exception as error:
                print(f"  ! GIF временно недоступен: {error}")
                return None
        return cache_path.read_bytes()

    print(f"  ! GIF не найден: {gif_value}")
    return None


def to_data_uri(data: bytes) -> str:
    return "data:image/gif;base64," + base64.b64encode(data).decode("ascii")


def load_icon_data_uri() -> str:
    icon = ROOT / "gifs" / "icon-180.png"
    if not icon.exists():
        return ""
    b64 = base64.b64encode(icon.read_bytes()).decode("ascii")
    return f"data:image/png;base64,{b64}"


def build_xlsx(programs: list[dict], out_path: Path, title: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as e:  # pragma: no cover
        raise RuntimeError(
            "Не могу собрать XLSX: нужен пакет openpyxl. Установи: pip install -r requirements.txt"
        ) from e

    wb = Workbook()
    wb.remove(wb.active)

    header = [
        "№",
        "Упражнение",
        "GIF",
        "Повторы/время",
        "Круги/подходы",
        "Отдых",
        "Цель",
        "Ключевая техника",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    accent = "9A4D76" if "девуш" in title.lower() else "176B87"
    header_fill = PatternFill("solid", fgColor=accent)
    wrap = Alignment(wrap_text=True, vertical="top")

    for p in programs:
        sheet_name = p.get("title") or p.get("id", "Программа")
        sheet_name = str(sheet_name)[:31]
        ws = wb.create_sheet(title=sheet_name)

        ws["A1"] = f'{title} · {p.get("title", "")}'
        ws["A1"].font = Font(bold=True, size=16, color=accent)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header))
        ws.row_dimensions[1].height = 28

        ws.append(header)
        for cell in ws[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")

        row = 3
        for s in p.get("sections", []):
            for ex in s.get("exercises", []):
                gif = ex.get("gif") or ""
                ws.cell(row=row, column=1, value=ex.get("num"))
                ws.cell(row=row, column=2, value=ex.get("name"))
                ws.cell(row=row, column=3, value=gif)
                if isinstance(gif, str) and gif.startswith(("http://", "https://")):
                    ws.cell(row=row, column=3).hyperlink = gif
                    ws.cell(row=row, column=3).style = "Hyperlink"
                ws.cell(row=row, column=4, value=ex.get("reps"))
                ws.cell(row=row, column=5, value=ex.get("rounds"))
                ws.cell(row=row, column=6, value=ex.get("rest"))
                ws.cell(row=row, column=7, value=ex.get("goal"))
                ws.cell(row=row, column=8, value=ex.get("technique"))

                for col in range(1, len(header) + 1):
                    ws.cell(row=row, column=col).alignment = wrap
                if row % 2:
                    for col in range(1, len(header) + 1):
                        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="F3F6F5")
                ws.row_dimensions[row].height = 42
                row += 1

        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:{get_column_letter(len(header))}{max(2, row - 1)}"

        widths = {1: 5, 2: 28, 3: 34, 4: 18, 5: 16, 6: 13, 7: 24, 8: 42}
        for c, w in widths.items():
            ws.column_dimensions[get_column_letter(c)].width = w

    wb.save(out_path)


def build_html(programs: list[dict]) -> str:
    icon_uri = load_icon_data_uri()
    icon_link = (
        f'<link rel="apple-touch-icon" href="{icon_uri}">\n'
        f'<link rel="icon" type="image/png" href="{icon_uri}">'
        if icon_uri
        else ""
    )

    def build_prog_group(group_programs: list[dict], group_id: str) -> tuple[list[str], list[str], list[str], list[str]]:
        tabs_inputs: list[str] = []
        tabs_labels: list[str] = []
        panels: list[str] = []
        show_rules: list[str] = []

        for i, p in enumerate(group_programs):
            checked = " checked" if i == 0 else ""
            tab_id = f'{group_id}-tab-{p["id"]}'
            panel_id = f'{group_id}-panel-{p["id"]}'
            tabs_inputs.append(
                f'<input class="tab-input" type="radio" name="{group_id}-prog" id="{tab_id}"{checked}>'
            )
            tabs_labels.append(
                f'<label class="tab" for="{tab_id}">{esc(p["badge"])}: {esc(p["title"].split()[0])}</label>'
            )

            sections_html = []
            for s in p["sections"]:
                items = []
                for ex in s["exercises"]:
                    chips = []
                    if ex.get("reps"):
                        chips.append(f'повторы: {ex["reps"]}')
                    if ex.get("rounds"):
                        chips.append(f'подходы: {ex["rounds"]}')
                    if ex.get("rest"):
                        chips.append(f'отдых: {ex["rest"]}')
                    if ex.get("equipment"):
                        chips.append(ex["equipment"])
                    chips_html = "".join(f'<span class="chip">{esc(c)}</span>' for c in chips)
                    gif_html = (
                        f'<div class="gif-frame"><img src="{esc(ex["gif"])}" '
                        f'alt="{esc(ex["name"])}" loading="lazy"></div>'
                        if ex.get("gif")
                        else '<div class="gif-frame gif-missing">Демонстрация скоро появится</div>'
                    )

                    items.append(
                        f"""
<details class="ex">
  <summary>
    <span class="ex-num">{esc(ex["num"])}</span>
    <span class="ex-text">
      <span class="ex-name">{esc(ex["name"])}</span>
      <span class="ex-meta">{esc(ex["reps"])} · {esc(ex["equipment"])}</span>
    </span>
    <span class="ex-hint">GIF ▼</span>
  </summary>
  <div class="ex-body">
    {gif_html}
    <div class="chips">{chips_html}</div>
    <p class="tech"><strong>Техника:</strong> {esc(ex.get("technique") or "")}</p>
    <p class="goal"><strong>Цель:</strong> {esc(ex.get("goal") or "")}</p>
  </div>
</details>"""
                    )

                sections_html.append(
                    f"""
<section class="section">
  <h3 class="section-title">{esc(s["title"])}</h3>
  <div class="list">{"".join(items)}</div>
</section>"""
                )

            panels.append(
                f"""
<div class="panel" id="{panel_id}">
  <div class="program-head">
    <h2>{esc(p["title"])}</h2>
    <p>{esc(p["subtitle"])}</p>
  </div>
  {"".join(sections_html)}
</div>"""
            )

            show_rules.append(
                f'#{tab_id}:checked ~ .tabs label[for="{tab_id}"] {{ background: var(--ink); color: #f7fbf8; }}'
            )
            show_rules.append(f'#{tab_id}:checked ~ #{panel_id} {{ display: block; }}')

        return tabs_inputs, tabs_labels, panels, show_rules

    female = [p for p in programs if (p.get("gender") or "unisex") in ("female", "unisex")]
    male = [p for p in programs if (p.get("gender") or "unisex") in ("male", "unisex")]

    g_inputs = [
        '<input class="tab-input" type="radio" name="gender" id="gender-f" checked>',
        '<input class="tab-input" type="radio" name="gender" id="gender-m">',
    ]
    g_tabs = [
        '<label class="tab tab-gender" for="gender-f">Девушка</label>',
        '<label class="tab tab-gender" for="gender-m">Парень</label>',
    ]

    f_inputs, f_labels, f_panels, f_rules = build_prog_group(female, "f")
    m_inputs, m_labels, m_panels, m_rules = build_prog_group(male, "m")

    show_rules = [
        '#gender-f:checked ~ .gender-tabs label[for="gender-f"] { background: var(--ink); color:#f7fbf8; }',
        '#gender-m:checked ~ .gender-tabs label[for="gender-m"] { background: var(--ink); color:#f7fbf8; }',
        '#gender-f:checked ~ .gender-panel-f { display:block; }',
        '#gender-m:checked ~ .gender-panel-m { display:block; }',
    ]
    show_rules.extend(f_rules)
    show_rules.extend(m_rules)

    return f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="default">
<meta name="apple-mobile-web-app-title" content="Гиря">
<meta name="mobile-web-app-capable" content="yes">
<title>Гиря</title>
<meta name="theme-color" content="#0b6e4f">
<link rel="manifest" href="manifest.webmanifest">
{icon_link}
<style>
:root {{
  --bg0:#dfe8e2; --bg1:#f3f7f4; --ink:#14201b; --muted:#5a6b63;
  --line:rgba(20,32,27,.12); --accent:#0b6e4f;
  --surface:rgba(255,255,255,.78); --shadow:0 14px 40px rgba(20,32,27,.12);
}}
* {{ box-sizing:border-box; }}
body {{
  margin:0;
  font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;
  color:var(--ink);
  background:
    radial-gradient(800px 420px at 8% -8%, rgba(11,110,79,.18), transparent 60%),
    radial-gradient(600px 360px at 100% 0%, rgba(240,162,2,.15), transparent 55%),
    linear-gradient(165deg,var(--bg0),var(--bg1) 50%, #e7eee9);
  min-height:100vh;
}}
.wrap {{ width:min(980px, calc(100% - 1.5rem)); margin:0 auto; padding:1rem 0 3rem; }}
.hero {{ display:grid; gap:.85rem; padding:1rem 0 1.4rem; }}
.brand {{
  margin:0; font-size:clamp(3.2rem, 14vw, 5.5rem); font-weight:800;
  letter-spacing:-.04em; line-height:.9;
  background:linear-gradient(120deg,var(--ink) 40%, var(--accent));
  -webkit-background-clip:text; background-clip:text; color:transparent;
}}
.hero h1 {{
  margin:0; font-size:clamp(1.15rem, 3.5vw, 1.55rem); font-weight:700; max-width:18ch; line-height:1.25;
}}
.hero p {{ margin:0; color:var(--muted); max-width:40ch; line-height:1.45; }}
.tip {{
  display:inline-block; margin-top:.2rem; padding:.55rem .9rem;
  border-radius:999px; background:rgba(11,110,79,.12); color:var(--accent);
  font-size:.9rem; font-weight:700;
}}
.programs {{ margin-top:.5rem; }}
.tab-input {{ position:absolute; opacity:0; pointer-events:none; }}
.gender-tabs {{
  display:grid; grid-template-columns:1fr 1fr; gap:.45rem; margin:0 0 1rem;
  padding:.4rem; border:1px solid var(--line); border-radius:22px;
  background:rgba(255,255,255,.62); box-shadow:0 10px 30px rgba(20,32,27,.08);
}}
.tabs {{
  display:flex; gap:.4rem; flex-wrap:wrap; margin:0 0 1rem;
  padding:.35rem; border:1px solid var(--line); border-radius:999px;
  background:rgba(243,247,244,.85);
}}
.tab {{
  padding:.7rem 1rem; border-radius:999px; cursor:pointer;
  color:var(--muted); font-weight:700; font-size:.92rem;
  -webkit-tap-highlight-color:transparent;
}}
.tab-gender {{ font-size:.95rem; }}
.gender-tabs .tab {{ text-align:center; padding:.85rem 1rem; }}
.gender-tabs label[for="gender-f"]::before {{ content:"♀"; margin-right:.4rem; }}
.gender-tabs label[for="gender-m"]::before {{ content:"♂"; margin-right:.4rem; }}
.gender-panel {{
  display:none; padding:1rem; border:1px solid var(--line); border-radius:24px;
  background:linear-gradient(145deg,var(--group-soft),rgba(255,255,255,.42));
}}
.gender-panel-f {{ --accent:#9a4d76; --group-soft:rgba(212,150,181,.18); }}
.gender-panel-m {{ --accent:#176b87; --group-soft:rgba(89,169,196,.16); }}
.panel {{ display:none; }}
{chr(10).join(show_rules)}
.group-label {{ margin:0 0 .8rem; font-size:1.1rem; font-weight:800; color:var(--accent); }}
.program-head h2 {{ margin:0 0 .3rem; font-size:clamp(1.25rem, 4vw, 1.8rem); }}
.program-head p {{ margin:0 0 1rem; color:var(--muted); }}
.section {{ margin-bottom:1.4rem; }}
.section-title {{
  margin:0 0 .7rem; color:var(--accent); text-transform:uppercase;
  letter-spacing:.04em; font-size:.85rem; font-weight:800;
}}
.list {{ display:grid; gap:.55rem; }}
.ex {{
  border:1px solid var(--line); border-radius:16px; background:var(--surface);
  overflow:hidden;
}}
.ex summary {{
  list-style:none; display:grid; grid-template-columns:auto 1fr auto;
  gap:.8rem; align-items:center; padding:.9rem 1rem; cursor:pointer;
  -webkit-tap-highlight-color:rgba(11,110,79,.15);
}}
.ex summary::-webkit-details-marker {{ display:none; }}
.ex[open] {{ box-shadow:var(--shadow); border-color:rgba(11,110,79,.35); }}
.ex[open] .ex-hint {{ color:var(--ink); }}
.ex-num {{
  width:2.1rem; height:2.1rem; border-radius:50%; display:grid; place-items:center;
  background:rgba(11,110,79,.12); color:var(--accent); font-weight:800;
}}
.ex-text {{ display:grid; gap:.15rem; min-width:0; }}
.ex-name {{ font-weight:700; font-size:1rem; }}
.ex-meta {{ color:var(--muted); font-size:.88rem; }}
.ex-hint {{ color:var(--accent); font-size:.78rem; font-weight:800; white-space:nowrap; }}
.ex-body {{ border-top:1px solid var(--line); }}
.gif-frame {{
  background:#101714; display:grid; place-items:center; min-height:220px;
}}
.gif-missing {{ color:#aebbb5; font-size:.9rem; }}
.gif-frame img {{
  width:100%; max-height:340px; object-fit:contain; display:block;
}}
.chips {{ display:flex; flex-wrap:wrap; gap:.35rem; padding:.9rem 1rem 0; }}
.chip {{
  font-size:.75rem; font-weight:700; padding:.3rem .6rem; border-radius:999px;
  background:rgba(11,110,79,.1); color:var(--accent);
}}
.tech, .goal {{
  margin:0; padding:.7rem 1rem 1rem; color:var(--muted); line-height:1.5; font-size:.95rem;
}}
.tech strong, .goal strong {{ color:var(--ink); }}
.goal {{ padding-top:0; }}
.footer-note {{ margin-top:1.5rem; color:var(--muted); font-size:.85rem; }}
@media (max-width:560px) {{
  .gender-panel {{ padding:.75rem; border-radius:20px; }}
  .ex summary {{ padding:.8rem; gap:.6rem; }}
}}
</style>
</head>
<body>
<div class="wrap">
  <header class="hero">
    <p class="brand">Гиря</p>
    <h1>Программа с гирей — техника в один тап</h1>
    <p>Нажми на упражнение — раскроется GIF с правильным выполнением и ключевыми нюансами.</p>
    <span class="tip">Жми на упражнение ↓</span>
  </header>

  <main class="programs" id="programs">
    {"".join(g_inputs)}
    <div class="gender-tabs">{"".join(g_tabs)}</div>

    <div class="gender-panel gender-panel-f">
      <p class="group-label">Программы для девушки</p>
      {"".join(f_inputs)}
      <div class="tabs">{"".join(f_labels)}</div>
      {"".join(f_panels)}
    </div>

    <div class="gender-panel gender-panel-m">
      <p class="group-label">Программы для парня</p>
      {"".join(m_inputs)}
      <div class="tabs">{"".join(m_labels)}</div>
      {"".join(m_panels)}
    </div>
    <p class="footer-note">После первой загрузки работает без интернета. На iPhone: Поделиться → На экран «Домой».</p>
  </main>
</div>
<script>
if ("serviceWorker" in navigator) {{
  window.addEventListener("load", function () {{
    navigator.serviceWorker.register("./sw.js").catch(function () {{}});
  }});
}}
</script>
</body>
</html>
"""


def write_offline_files(cache_version: str) -> None:
    # иконку для манифеста не обязательна — apple-touch-icon уже в HTML
    manifest = {
        "name": "Гиря",
        "short_name": "Гиря",
        "description": "Программы с гирей офлайн",
        "start_url": "./",
        "scope": "./",
        "display": "standalone",
        "background_color": "#e7eee9",
        "theme_color": "#0b6e4f",
        "lang": "ru",
    }
    (ROOT / "manifest.webmanifest").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )

    sw = f"""/* Offline cache for Гиря */
const CACHE = "girya-{cache_version}";
const ASSETS = ["./", "./index.html", "./manifest.webmanifest"];

self.addEventListener("install", (event) => {{
  event.waitUntil(
    caches.open(CACHE).then((cache) => cache.addAll(ASSETS)).then(() => self.skipWaiting())
  );
}});

self.addEventListener("activate", (event) => {{
  event.waitUntil(
    caches.keys().then((keys) =>
      Promise.all(keys.filter((k) => k !== CACHE).map((k) => caches.delete(k)))
    ).then(() => self.clients.claim())
  );
}});

self.addEventListener("fetch", (event) => {{
  if (event.request.method !== "GET") return;
  event.respondWith(
    caches.match(event.request).then((cached) => {{
      if (cached) return cached;
      return fetch(event.request).then((response) => {{
        const copy = response.clone();
        if (response.ok && new URL(event.request.url).origin === self.location.origin) {{
          caches.open(CACHE).then((cache) => cache.put(event.request, copy));
        }}
        return response;
      }}).catch(() => caches.match("./index.html"));
    }})
  );
}});
"""
    (ROOT / "sw.js").write_text(sw, encoding="utf-8")


def load_programs() -> list[dict]:
    raw = json.loads(PROGRAMS_FILE.read_text(encoding="utf-8"))
    if isinstance(raw, dict):
        programs = raw.get("programs")
        if not isinstance(programs, list):
            raise ValueError('В programs.json нужен массив "programs"')
        return programs
    if isinstance(raw, list):
        return raw
    raise ValueError("programs.json должен быть объектом {programs:[...]} или массивом")


def main() -> None:
    programs = load_programs()

    female = [p for p in programs if p.get("gender") == "female"]
    male = [p for p in programs if p.get("gender") == "male"]

    print("Собираю XLSX из programs.json...")
    build_xlsx(female, OUT_XLSX_FEMALE, "Программа тренировок (девушка)")
    build_xlsx(male, OUT_XLSX_MALE, "Программа тренировок (парень)")
    print(f"  • {OUT_XLSX_FEMALE.name}")
    print(f"  • {OUT_XLSX_MALE.name}")

    print("Вшиваю GIF и собираю HTML...")
    for p in programs:
        for s in p["sections"]:
            for ex in s["exercises"]:
                name = ex["name"]
                src = ex.get("gif") or ""
                print(f"  • {name}")
                data = load_gif_bytes(src)
                ex["gif"] = to_data_uri(data) if data else ""

    html = build_html(programs)
    OUT_HTML.write_text(html, encoding="utf-8")
    (ROOT / "index.html").write_text(html, encoding="utf-8")

    # версия кэша меняется при каждой сборке — iPhone подтянет обновление
    cache_version = str(int((ROOT / "index.html").stat().st_mtime))
    write_offline_files(cache_version)

    with zipfile.ZipFile(OUT_ZIP, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.write(OUT_HTML, arcname="Гиря.html")
        zf.write(OUT_XLSX_FEMALE, arcname=OUT_XLSX_FEMALE.name)
        zf.write(OUT_XLSX_MALE, arcname=OUT_XLSX_MALE.name)

    size_mb = OUT_HTML.stat().st_size / 1024 / 1024
    print(f"\nГотово: {OUT_HTML.name} ({size_mb:.1f} МБ)")
    print(f"Архив: {OUT_ZIP.name}")
    print("Офлайн: sw.js + manifest.webmanifest")


if __name__ == "__main__":
    main()
